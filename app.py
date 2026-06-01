from pathlib import Path
import sqlite3
import time
from datetime import datetime
from io import BytesIO
from threading import Lock

from flask import Flask, Response, render_template, request
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import Title
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, RegularTextRun
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DB_PATH = Path("/data/NetikotDB.db")
DB_URI = f"file:{DB_PATH.as_posix()}?mode=ro&immutable=1"
ROW_LIMIT = 500
DASHBOARD_CACHE_TTL_SECONDS = 5 * 60

app = Flask(__name__)

_dashboard_cache = {
    "expires_at": 0,
    "html": None,
}
_dashboard_cache_lock = Lock()

LATEST_RUN_CTE = """
WITH latest_run AS (
    SELECT results.*
    FROM tbl_Results AS results
    JOIN (
        SELECT IP_Address, MAX("Date") AS latest_date
        FROM tbl_Results
        GROUP BY IP_Address
    ) AS latest
        ON latest.IP_Address = results.IP_Address
        AND latest.latest_date = results."Date"
)
"""

LAST_EXECUTION_CTE = """
WITH ordered_results AS (
    SELECT
        results.*,
        LAG(datetime("Date")) OVER (ORDER BY datetime("Date"), ID) AS previous_date
    FROM tbl_Results AS results
),
run_starts AS (
    SELECT ID
    FROM ordered_results
    WHERE previous_date IS NULL
        OR (julianday("Date") - julianday(previous_date)) * 24 * 60 >= 30
),
last_execution AS (
    SELECT results.*
    FROM tbl_Results AS results
    WHERE results.ID >= (SELECT MAX(ID) FROM run_starts)
)
"""

FAILURE_TYPE_EXPR = """
CASE
    WHEN LOWER(TRIM(COALESCE(Details, ''))) LIKE '%connection aborted%' THEN 'Connection aborted'
    WHEN LOWER(TRIM(COALESCE(Details, ''))) LIKE '%timeout%' THEN 'timeout'
    ELSE COALESCE(NULLIF(TRIM(Details), ''), 'No details')
END
"""


def get_connection():
    connection = sqlite3.connect(DB_URI, uri=True)
    connection.row_factory = sqlite3.Row
    return connection


def query_all(query, params=()):
    with get_connection() as connection:
        return connection.execute(query, params).fetchall()


def query_one(query, params=()):
    with get_connection() as connection:
        return connection.execute(query, params).fetchone()


def query_all_with_fallback(query, fallback_query, params=()):
    try:
        return query_all(query, params)
    except sqlite3.OperationalError as error:
        message = str(error).lower()
        if "no such table" not in message and "no such view" not in message:
            raise
        return query_all(fallback_query, params)


def get_cached_dashboard_html():
    if _dashboard_cache["html"] and time.monotonic() < _dashboard_cache["expires_at"]:
        return _dashboard_cache["html"]
    return None


def set_cached_dashboard_html(html):
    _dashboard_cache["html"] = html
    _dashboard_cache["expires_at"] = time.monotonic() + DASHBOARD_CACHE_TTL_SECONDS


def format_file_size(size_in_bytes):
    size = float(size_in_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} {unit}"
        size /= 1024


def build_failure_reason_chart(days, rows):
    day_list = [row["day"] for row in days if row["day"]]
    if not day_list:
        return {
            "days": [],
            "labels": [],
            "reasons": [],
            "series": [],
            "total_series": None,
            "max_amount": 0,
            "y_labels": [],
            "date_range": "No data",
        }

    totals = {}
    amounts = {}
    day_totals = {day: 0 for day in day_list}
    for row in rows:
        day = row["day"]
        reason = row["reason"] or "No details"
        amount = row["amount"] or 0
        totals[reason] = totals.get(reason, 0) + amount
        amounts[(reason, day)] = amount
        if day in day_totals:
            day_totals[day] += amount

    max_amount = max([row["amount"] or 0 for row in rows] or [0])
    max_total_amount = max(day_totals.values() or [0])
    chart_max = max(max_amount, max_total_amount, 1)
    x_start = 6
    x_width = 88
    y_top = 12
    y_bottom = 86
    y_height = y_bottom - y_top
    day_denominator = max(len(day_list) - 1, 1)

    reasons = []
    series = []
    total_dots = []
    for day_index, day in enumerate(day_list):
        x = 50 if len(day_list) == 1 else x_start + (x_width * day_index / day_denominator)
        amount = day_totals.get(day, 0)
        y = y_bottom - (y_height * amount / chart_max)
        total_dots.append(
            {
                "day": day,
                "amount": amount,
                "x": round(x, 2),
                "y": round(y, 2),
            }
        )

    total_series = {
        "reason": "Total failures",
        "total": sum(day_totals.values()),
        "points": " ".join(f"{dot['x']},{dot['y']}" for dot in total_dots),
        "dots": total_dots,
    }

    for reason in sorted(totals, key=lambda item: (-totals[item], item.lower())):
        cells = [
            {
                "day": day,
                "amount": amounts.get((reason, day), 0),
            }
            for day in day_list
        ]
        dots = []
        for day_index, cell in enumerate(cells):
            x = 50 if len(day_list) == 1 else x_start + (x_width * day_index / day_denominator)
            y = y_bottom - (y_height * cell["amount"] / chart_max)
            dots.append(
                {
                    "day": cell["day"],
                    "amount": cell["amount"],
                    "x": round(x, 2),
                    "y": round(y, 2),
                }
            )

        reasons.append(
            {
                "reason": reason,
                "total": totals[reason],
                "cells": cells,
            }
        )
        series.append(
            {
                "reason": reason,
                "total": totals[reason],
                "points": " ".join(f"{dot['x']},{dot['y']}" for dot in dots),
                "dots": dots,
            }
        )

    y_labels = [
        round(chart_max),
        round(chart_max * 0.75),
        round(chart_max * 0.5),
        round(chart_max * 0.25),
        0,
    ]

    return {
        "days": day_list,
        "labels": [day[5:] for day in day_list],
        "reasons": reasons,
        "series": series,
        "total_series": total_series,
        "max_amount": max_amount,
        "y_labels": y_labels,
        "date_range": f"{day_list[0]} - {day_list[-1]}",
    }


def get_failure_reason_days():
    return query_all_with_fallback(
        """
        SELECT DISTINCT date("Date") AS day
        FROM vew_LastMonthRun
        WHERE "Date" IS NOT NULL
        ORDER BY day;
        """,
        """
        SELECT DISTINCT date("Date") AS day
        FROM tbl_Results
        WHERE "Date" >= datetime((SELECT MAX("Date") FROM tbl_Results), '-30 days')
        ORDER BY day;
        """,
    )


def get_failure_reason_summary_rows():
    return query_all_with_fallback(
        f"""
        WITH ranked AS (
            SELECT
                IP_Address,
                Is_Success,
                {FAILURE_TYPE_EXPR} AS reason,
                date("Date") AS day,
                ROW_NUMBER() OVER (
                    PARTITION BY IP_Address, date("Date")
                    ORDER BY datetime("Date") DESC
                ) AS row_number
            FROM vew_LastMonthRun
            WHERE "Date" IS NOT NULL
        )
        SELECT day, reason, COUNT(*) AS amount
        FROM ranked
        WHERE row_number = 1 AND Is_Success = 0
        GROUP BY day, reason
        ORDER BY day, amount DESC, reason;
        """,
        f"""
        WITH ranked AS (
            SELECT
                IP_Address,
                Is_Success,
                {FAILURE_TYPE_EXPR} AS reason,
                date("Date") AS day,
                ROW_NUMBER() OVER (
                    PARTITION BY IP_Address, date("Date")
                    ORDER BY datetime("Date") DESC
                ) AS row_number
            FROM tbl_Results
            WHERE "Date" >= datetime((SELECT MAX("Date") FROM tbl_Results), '-30 days')
        )
        SELECT day, reason, COUNT(*) AS amount
        FROM ranked
        WHERE row_number = 1 AND Is_Success = 0
        GROUP BY day, reason
        ORDER BY day, amount DESC, reason;
        """,
    )


def group_counts_by_day(rows):
    groups = []
    group_lookup = {}

    for row in rows:
        day = row["day"] or "No date"
        if day not in group_lookup:
            day_group = {
                "day": day,
                "total": 0,
                "reasons": [],
            }
            group_lookup[day] = day_group
            groups.append(day_group)
        else:
            day_group = group_lookup[day]

        amount = row["amount"] or 0
        day_group["reasons"].append(
            {
                "reason": row["reason"] or "No details",
                "amount": amount,
            }
        )
        day_group["total"] += amount

    return sorted(groups, key=lambda group: group["day"], reverse=True)


def get_failure_reason_chart(days=None, rows=None):
    return build_failure_reason_chart(
        days if days is not None else get_failure_reason_days(),
        rows if rows is not None else get_failure_reason_summary_rows(),
    )


def get_latest_failure_type_counts(limit, detail_alias="detail", amount_alias="amount"):
    return query_all(
        LATEST_RUN_CTE
        + f"""
        SELECT
            {FAILURE_TYPE_EXPR} AS "{detail_alias}",
            COUNT(*) AS "{amount_alias}"
        FROM latest_run
        WHERE Is_Success = 0
        GROUP BY {FAILURE_TYPE_EXPR}
        ORDER BY "{amount_alias}" DESC, "{detail_alias}"
        LIMIT ?;
        """,
        (limit,),
    )


def get_incorrect_password_rows(limit):
    return query_all(
        LATEST_RUN_CTE
        + """
        SELECT IP_Address, Site_Name, Details, "Date"
        FROM latest_run
        WHERE LOWER(COALESCE(Details, '')) LIKE '%password is incorrect%'
        ORDER BY "Date" DESC
        LIMIT ?;
        """,
        (limit,),
    )


def get_port_closed_rows(limit):
    return query_all(
        LATEST_RUN_CTE
        + """
        SELECT IP_Address, Site_Name, Details, "Date"
        FROM latest_run
        WHERE LOWER(TRIM(COALESCE(Details, ''))) = 'port is closed'
        ORDER BY "Date" DESC
        LIMIT ?;
        """,
        (limit,),
    )


def get_ip_errors_since_success(limit):
    return query_all(
        LATEST_RUN_CTE
        + """
        , current_failed AS (
            SELECT
                IP_Address,
                Site_Name,
                COALESCE(NULLIF(TRIM(Details), ''), 'No details') AS Last_error_type,
                "Date" AS failed_date
            FROM latest_run
            WHERE Is_Success = 0
        ),
        last_success AS (
            SELECT
                current_failed.IP_Address,
                MAX(results."Date") AS last_success_date
            FROM current_failed
            LEFT JOIN tbl_Results AS results
                ON results.IP_Address = current_failed.IP_Address
                AND results.Is_Success = 1
                AND datetime(results."Date") < datetime(current_failed.failed_date)
            GROUP BY current_failed.IP_Address
        ),
        counted AS (
            SELECT
                current_failed.Site_Name,
                current_failed.IP_Address,
                current_failed.Last_error_type,
                current_failed.failed_date,
                last_success.last_success_date,
                COUNT(results.ID) AS Errors
            FROM current_failed
            LEFT JOIN last_success ON last_success.IP_Address = current_failed.IP_Address
            LEFT JOIN tbl_Results AS results
                ON results.IP_Address = current_failed.IP_Address
                AND results.Is_Success = 0
                AND datetime(results."Date") <= datetime(current_failed.failed_date)
                AND (
                    last_success.last_success_date IS NULL
                    OR datetime(results."Date") > datetime(last_success.last_success_date)
                )
            GROUP BY
                current_failed.IP_Address,
                current_failed.Site_Name,
                current_failed.Last_error_type,
                current_failed.failed_date,
                last_success.last_success_date
        )
        SELECT
            IP_Address,
            Site_Name,
            Errors,
            Last_error_type,
            COALESCE(date(last_success_date), 'No success') AS Last_success_date
        FROM counted
        ORDER BY Errors DESC, datetime(failed_date) DESC, IP_Address
        LIMIT ?;
        """,
        (limit,),
    )


def build_ip_error_distribution(rows):
    buckets = {}
    for row in rows:
        errors = row["Errors"] or 0
        buckets[errors] = buckets.get(errors, 0) + 1

    max_ips = max(buckets.values() or [1])
    return [
        {
            "errors": errors,
            "ip_count": ip_count,
            "height": round(ip_count / max_ips * 100, 1),
        }
        for errors, ip_count in sorted(buckets.items())
    ]


def build_ip_error_y_ticks(chart_rows):
    max_ips = max([row["ip_count"] for row in chart_rows] or [0])
    if max_ips <= 0:
        return []

    raw_ticks = [0, round(max_ips * 0.25), round(max_ips * 0.5), round(max_ips * 0.75), max_ips]
    ticks = sorted(set(raw_ticks))
    return [
        {
            "value": tick,
            "position": round(tick / max_ips * 100, 1),
        }
        for tick in ticks
    ]


def render_rows(title, rows, description=None):
    return render_template(
        "results.html",
        title=title,
        rows=rows,
        description=description,
        row_limit=ROW_LIMIT,
    )


def get_weekly_failure_reason_days():
    return query_all_with_fallback(
        """
        SELECT DISTINCT date("Date") AS day
        FROM vew_LastMonthRun
        WHERE "Date" IS NOT NULL
            AND date("Date") >= date((SELECT MAX("Date") FROM vew_LastMonthRun), '-6 days')
        ORDER BY day;
        """,
        """
        SELECT DISTINCT date("Date") AS day
        FROM tbl_Results
        WHERE "Date" IS NOT NULL
            AND date("Date") >= date((SELECT MAX("Date") FROM tbl_Results), '-6 days')
        ORDER BY day;
        """,
    )


def get_weekly_failure_reason_summary_rows():
    return query_all_with_fallback(
        """
        WITH ranked AS (
            SELECT
                IP_Address,
                Is_Success,
                COALESCE(NULLIF(TRIM(Details), ''), 'No details') AS reason,
                date("Date") AS day,
                ROW_NUMBER() OVER (
                    PARTITION BY IP_Address, date("Date")
                    ORDER BY datetime("Date") DESC
                ) AS row_number
            FROM vew_LastMonthRun
            WHERE "Date" IS NOT NULL
                AND date("Date") >= date((SELECT MAX("Date") FROM vew_LastMonthRun), '-6 days')
        )
        SELECT day, reason, COUNT(*) AS amount
        FROM ranked
        WHERE row_number = 1 AND Is_Success = 0
        GROUP BY day, reason
        ORDER BY day, amount DESC, reason;
        """,
        """
        WITH ranked AS (
            SELECT
                IP_Address,
                Is_Success,
                COALESCE(NULLIF(TRIM(Details), ''), 'No details') AS reason,
                date("Date") AS day,
                ROW_NUMBER() OVER (
                    PARTITION BY IP_Address, date("Date")
                    ORDER BY datetime("Date") DESC
                ) AS row_number
            FROM tbl_Results
            WHERE "Date" IS NOT NULL
                AND date("Date") >= date((SELECT MAX("Date") FROM tbl_Results), '-6 days')
        )
        SELECT day, reason, COUNT(*) AS amount
        FROM ranked
        WHERE row_number = 1 AND Is_Success = 0
        GROUP BY day, reason
        ORDER BY day, amount DESC, reason;
        """,
    )


def build_weekly_failure_types_xlsx():
    days = [row["day"] for row in get_weekly_failure_reason_days() if row["day"]]
    rows = list(get_weekly_failure_reason_summary_rows())
    chart = build_failure_reason_chart([{"day": day} for day in days], rows)

    workbook = Workbook()
    worksheet = workbook.active
    chart_sheet = worksheet
    chart_sheet.title = "Weekly Report"
    list_sheet = workbook.create_sheet("Weekly Failure Types")
    ip_graph_sheet = workbook.create_sheet("IP Errors Graph")
    ip_table_sheet = workbook.create_sheet("IP Errors Table")
    data_sheet = workbook.create_sheet("Chart Data")
    data_sheet.sheet_state = "hidden"
    ip_chart_data_sheet = workbook.create_sheet("IP Error Chart Data")
    ip_chart_data_sheet.sheet_state = "hidden"

    header_fill = PatternFill("solid", fgColor="315BE8")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="0B1739", bold=True, size=16)
    section_font = Font(color="0B1739", bold=True, size=12)
    muted_font = Font(color="64718B")
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    def shorten_chart_label(value, limit=28):
        text = str(value or "No details")
        return text if len(text) <= limit else f"{text[:limit - 3]}..."

    def chart_text(value, size=1000, bold=False):
        return Text(
            rich=RichText(
                p=[
                    Paragraph(
                        r=[
                            RegularTextRun(
                                rPr=CharacterProperties(sz=size, b=bold),
                                t=value,
                            )
                        ]
                    )
                ]
            )
        )

    def chart_text_title(value, x, y, width, height, size=1000, bold=False):
        return Title(
            tx=chart_text(value, size=size, bold=bold),
            layout=Layout(
                manualLayout=ManualLayout(
                    xMode="factor",
                    yMode="factor",
                    wMode="factor",
                    hMode="factor",
                    x=x,
                    y=y,
                    w=width,
                    h=height,
                )
            ),
            overlay=True,
        )

    chart_sheet.merge_cells("A1:S1")
    chart_sheet.merge_cells("A2:S2")
    chart_sheet["A1"] = "Weekly Report"
    chart_sheet["A1"].font = title_font
    chart_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    chart_sheet["A2"] = chart["date_range"]
    chart_sheet["A2"].font = muted_font
    chart_sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    chart_sheet.row_dimensions[1].height = 28
    chart_sheet.row_dimensions[2].height = 22
    if not days:
        chart_sheet["A4"] = "No weekly failure types found."
        list_sheet["A1"] = "No weekly failure types found."
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    reasons = [reason["reason"] for reason in chart["reasons"]]
    chart_header = ["Date", "Total failures", *[shorten_chart_label(reason) for reason in reasons]]
    for column_index, value in enumerate(chart_header, start=1):
        cell = data_sheet.cell(row=1, column=column_index, value=value)
        cell.font = header_font
        cell.fill = header_fill

    day_totals = {point["day"]: point["amount"] for point in chart["total_series"]["dots"]}
    reason_amounts = {
        (series["reason"], point["day"]): point["amount"]
        for series in chart["series"]
        for point in series["dots"]
    }

    for row_index, day in enumerate(days, start=2):
        data_sheet.cell(row=row_index, column=1, value=day)
        data_sheet.cell(row=row_index, column=2, value=day_totals.get(day, 0))
        for column_index, reason in enumerate(reasons, start=3):
            data_sheet.cell(row=row_index, column=column_index, value=reason_amounts.get((reason, day), 0))

    chart_end_row = 1 + len(days)
    if chart_end_row >= 2:
        line_chart = LineChart()
        line_chart.title = Title(tx=chart_text("Weekly Failure Types", size=1400, bold=True))
        line_chart.style = 2
        line_chart.y_axis.title = Title(tx=chart_text("Failures", size=900))
        line_chart.x_axis.title = chart_text_title("Date", 0.12, 0.86, 0.08, 0.04, size=900)
        line_chart.height = 10.6
        line_chart.width = 23.2
        line_chart.legend.position = "r"
        line_chart.legend.overlay = False
        line_chart.layout = Layout(
            manualLayout=ManualLayout(
                xMode="factor",
                yMode="factor",
                wMode="factor",
                hMode="factor",
                x=0.0,
                y=0.02,
                w=0.78,
                h=0.82,
            )
        )
        line_chart.y_axis.scaling.min = 0
        line_chart.y_axis.majorGridlines.spPr = GraphicalProperties(
            ln=LineProperties(solidFill="D9D2C3", w=8000)
        )
        line_chart.graphical_properties = GraphicalProperties(
            ln=LineProperties(solidFill="9E9E9E", w=8000)
        )
        data = Reference(data_sheet, min_col=2, max_col=max(2, len(chart_header)), min_row=1, max_row=chart_end_row)
        categories = Reference(data_sheet, min_col=1, min_row=2, max_row=chart_end_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        palette = ["B65A42", "3D5F91", "6D8B3A", "C68143", "5C78B2", "4F8E87", "9B6BB8", "D88B8B"]
        marker_symbols = ["circle", "diamond", "circle", "diamond", "circle", "diamond", "circle", "diamond"]
        for index, series in enumerate(line_chart.series):
            color = palette[index % len(palette)]
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 18000
            series.marker.symbol = marker_symbols[index % len(marker_symbols)]
            series.marker.size = 6
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = color
        chart_sheet.add_chart(line_chart, "A3")

    rows_by_day = {}
    for row in rows:
        day = row["day"] or "No date"
        rows_by_day.setdefault(day, []).append(row)

    current_row = 1
    for day in sorted(rows_by_day, reverse=True):
        day_rows = sorted(
            rows_by_day[day],
            key=lambda row: (-(row["amount"] or 0), row["reason"] or ""),
        )
        day_total = sum(row["amount"] or 0 for row in day_rows)

        list_sheet.cell(row=current_row, column=1, value=day)
        list_sheet.cell(row=current_row, column=2, value="Total failures")
        list_sheet.cell(row=current_row, column=3, value=day_total)
        for column_index in range(1, 4):
            cell = list_sheet.cell(row=current_row, column=column_index)
            cell.font = section_font
            cell.border = thin_border
        current_row += 1

        for column_index, value in enumerate(["Failure type", "Amount"], start=1):
            cell = list_sheet.cell(row=current_row, column=column_index, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for row in day_rows:
            list_sheet.cell(row=current_row, column=1, value=row["reason"] or "No details")
            list_sheet.cell(row=current_row, column=2, value=row["amount"] or 0)
            current_row += 1

        current_row += 1

    for row in list_sheet.iter_rows(min_row=1, max_row=list_sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical="top", wrap_text=False)

    widths = {
        1: 16,
        2: 72,
        3: 14,
    }
    for column_index in range(1, 10):
        width = widths.get(column_index, 18)
        list_sheet.column_dimensions[get_column_letter(column_index)].width = width
    for column_index in range(1, 24):
        chart_sheet.column_dimensions[get_column_letter(column_index)].width = 14

    ip_error_rows = list(get_ip_errors_since_success(ROW_LIMIT))
    ip_chart_rows = ip_error_rows[:12]

    ip_table_headers = ["IP", "Site", "Errors", "Last type", "Last success date"]
    for column_index, value in enumerate(ip_table_headers, start=1):
        cell = ip_table_sheet.cell(row=1, column=column_index, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(ip_error_rows, start=2):
        ip_table_sheet.cell(row=row_index, column=1, value=row["IP_Address"])
        ip_table_sheet.cell(row=row_index, column=2, value=row["Site_Name"])
        ip_table_sheet.cell(row=row_index, column=3, value=row["Errors"] or 0)
        ip_table_sheet.cell(row=row_index, column=4, value=row["Last_error_type"] or "No details")
        ip_table_sheet.cell(row=row_index, column=5, value=row["Last_success_date"])

    for row in ip_table_sheet.iter_rows(min_row=1, max_row=ip_table_sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical="top", wrap_text=False)

    ip_table_widths = {
        1: 18,
        2: 34,
        3: 12,
        4: 46,
        5: 16,
    }
    for column_index, width in ip_table_widths.items():
        ip_table_sheet.column_dimensions[get_column_letter(column_index)].width = width
    if ip_error_rows:
        ip_table_sheet.auto_filter.ref = f"A1:E{ip_table_sheet.max_row}"

    ip_chart_data_sheet.cell(row=1, column=1, value="IP")
    ip_chart_data_sheet.cell(row=1, column=2, value="Errors")
    for row_index, row in enumerate(ip_chart_rows, start=2):
        ip_chart_data_sheet.cell(row=row_index, column=1, value=row["IP_Address"])
        ip_chart_data_sheet.cell(row=row_index, column=2, value=row["Errors"] or 0)

    ip_graph_sheet.merge_cells("A1:K1")
    ip_graph_sheet["A1"] = "Most errors per IP since last success"
    ip_graph_sheet["A1"].font = title_font
    ip_graph_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ip_graph_sheet.row_dimensions[1].height = 28
    for column_index in range(1, 14):
        ip_graph_sheet.column_dimensions[get_column_letter(column_index)].width = 14

    if ip_chart_rows:
        def make_ip_bar_chart(width, height):
            ip_bar_chart = BarChart()
            ip_bar_chart.type = "col"
            ip_bar_chart.title = Title(tx=chart_text("Most Errors Per IP Since Last Success", size=1400, bold=True))
            ip_bar_chart.style = 10
            ip_bar_chart.y_axis.title = Title(tx=chart_text("Errors", size=900))
            ip_bar_chart.x_axis.title = None
            ip_bar_chart.legend = None
            ip_bar_chart.height = height
            ip_bar_chart.width = width
            ip_bar_chart.y_axis.scaling.min = 0
            ip_bar_chart.y_axis.majorGridlines.spPr = GraphicalProperties(
                ln=LineProperties(solidFill="D9D2C3", w=8000)
            )
            ip_bar_chart.graphical_properties = GraphicalProperties(
                ln=LineProperties(solidFill="9E9E9E", w=8000)
            )
            data = Reference(ip_chart_data_sheet, min_col=2, min_row=1, max_row=len(ip_chart_rows) + 1)
            categories = Reference(ip_chart_data_sheet, min_col=1, min_row=2, max_row=len(ip_chart_rows) + 1)
            ip_bar_chart.add_data(data, titles_from_data=True)
            ip_bar_chart.set_categories(categories)
            if ip_bar_chart.series:
                ip_bar_chart.series[0].graphicalProperties.solidFill = "5C78B2"
                ip_bar_chart.series[0].graphicalProperties.line.solidFill = "5C78B2"
            return ip_bar_chart

        ip_graph_sheet.add_chart(make_ip_bar_chart(18.8, 12), "A3")
        chart_sheet.add_chart(make_ip_bar_chart(21.0, 10.6), "J3")
    else:
        ip_graph_sheet["A3"] = "No IP errors since last success found."

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


@app.route("/download_report")
def download_report():
    filename = f"weekly_report-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return Response(
        build_weekly_failure_types_xlsx(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/")
def index():
    cached_html = get_cached_dashboard_html()
    if cached_html is not None:
        return cached_html

    with _dashboard_cache_lock:
        cached_html = get_cached_dashboard_html()
        if cached_html is not None:
            return cached_html

        html = build_dashboard_html()
        set_cached_dashboard_html(html)
        return html


def build_dashboard_html():
    db_size_label = format_file_size(DB_PATH.stat().st_size) if DB_PATH.exists() else "Missing"
    today_date = datetime.now().strftime("%Y-%m-%d")

    summary = query_one(
        """
        SELECT
            COUNT(*) AS total_checks,
            COALESCE(SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END), 0) AS success_checks,
            COALESCE(SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END), 0) AS failed_checks,
            COUNT(DISTINCT IP_Address) AS unique_ips,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END) /
                NULLIF(COUNT(*), 0),
                1
            ), 0) AS success_rate,
            MIN("Date") AS first_seen,
            MAX("Date") AS last_seen
        FROM tbl_Results;
        """
    )

    target_summary = query_one(
        LATEST_RUN_CTE
        + """
        SELECT
            COUNT(*) AS total_targets,
            COALESCE(SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END), 0) AS healthy_targets,
            COALESCE(SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END), 0) AS failing_targets,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END) /
                NULLIF(COUNT(*), 0),
                1
            ), 0) AS healthy_rate
        FROM latest_run;
        """
    )

    last_execution_summary = query_one(
        LAST_EXECUTION_CTE
        + """
        SELECT
            COUNT(*) AS total,
            date(MAX("Date")) AS run_date,
            COALESCE(SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END), 0) AS success_total,
            COALESCE(SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END), 0) AS failed_total,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END) / NULLIF(COUNT(*), 0),
                1
            ), 0) AS success_rate,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END) / NULLIF(COUNT(*), 0),
                1
            ), 0) AS failed_rate
        FROM last_execution;
        """,
    )

    last_run_card = {
        "total": last_execution_summary["total"],
        "run_date": last_execution_summary["run_date"],
    }
    last_success_card = {
        "total": last_execution_summary["success_total"],
        "run_date": last_execution_summary["run_date"],
        "rate": last_execution_summary["success_rate"],
    }
    last_failed_card = {
        "total": last_execution_summary["failed_total"],
        "run_date": last_execution_summary["run_date"],
        "rate": last_execution_summary["failed_rate"],
    }

    top_failure_details = get_latest_failure_type_counts(10)

    site_rows = query_all(
        LATEST_RUN_CTE
        + """
        SELECT
            Site_Name,
            COUNT(*) AS total,
            COALESCE(SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END), 0) AS success,
            COALESCE(SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END), 0) AS failed,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END) /
                NULLIF(COUNT(*), 0),
                1
            ), 0) AS success_rate
        FROM latest_run
        GROUP BY Site_Name
        ORDER BY failed DESC, total DESC
        LIMIT 10;
        """
    )

    incorrect_passwords = get_incorrect_password_rows(10)
    port_closed_rows = get_port_closed_rows(10)

    latest_success_rows = query_all(
        LATEST_RUN_CTE
        + """
        SELECT IP_Address, Site_Name, date("Date") AS "Date"
        FROM latest_run
        WHERE Is_Success = 1
        ORDER BY "Date" DESC
        LIMIT 10;
        """
    )
    ip_error_chart_source = get_ip_errors_since_success(ROW_LIMIT)
    ip_error_chart_rows = build_ip_error_distribution(ip_error_chart_source)
    ip_error_chart_y_ticks = build_ip_error_y_ticks(ip_error_chart_rows)
    ip_error_rows = ip_error_chart_source[:10]

    max_failure_amount = max([row["amount"] for row in top_failure_details] or [1])
    failure_reason_chart = get_failure_reason_chart()

    return render_template(
        "index.html",
        db_size_label=db_size_label,
        summary=summary,
        target_summary=target_summary,
        last_run_card=last_run_card,
        today_date=today_date,
        last_success_card=last_success_card,
        last_failed_card=last_failed_card,
        failure_reason_chart=failure_reason_chart,
        top_failure_details=top_failure_details,
        site_rows=site_rows,
        incorrect_passwords=incorrect_passwords,
        port_closed_rows=port_closed_rows,
        latest_success_rows=latest_success_rows,
        ip_error_rows=ip_error_rows,
        ip_error_chart_rows=ip_error_chart_rows,
        ip_error_chart_y_ticks=ip_error_chart_y_ticks,
        max_failure_amount=max_failure_amount,
    )


@app.route("/monthly_failure_types")
def monthly_failure_types():
    rows = get_failure_reason_summary_rows()
    chart = get_failure_reason_chart(rows=rows)
    grouped_counts = group_counts_by_day(rows)
    return render_template(
        "failure_reasons.html",
        failure_reason_chart=chart,
        rows=rows,
        grouped_counts=grouped_counts,
    )


@app.route("/failure_reasons")
def failure_reasons():
    return monthly_failure_types()


@app.route("/latest_failures")
def latest_failures():
    rows = query_all(
        LATEST_RUN_CTE
        + """
        SELECT ID, IP_Address, Port, Site_Name, Details, "Date"
        FROM latest_run
        WHERE Is_Success = 0
        ORDER BY "Date" DESC
        LIMIT ?;
        """,
        (ROW_LIMIT,),
    )
    return render_rows("Latest failures", rows, "Latest failed result rows.")


@app.route("/incorrect_passwords")
def incorrect_passwords():
    rows = get_incorrect_password_rows(ROW_LIMIT)
    return render_rows("Incorrect password targets", rows, "Password errors from the latest run.")


@app.route("/port_closed")
def port_closed():
    rows = get_port_closed_rows(ROW_LIMIT)
    return render_rows("Port closed targets", rows, "Latest targets with Port is closed failures.")


@app.route("/latest_successes")
def latest_successes():
    rows = query_all(
        LATEST_RUN_CTE
        + """
        SELECT ID, IP_Address, Port, Site_Name, Details, "Date"
        FROM latest_run
        WHERE Is_Success = 1
        ORDER BY "Date" DESC
        LIMIT ?;
        """,
        (ROW_LIMIT,),
    )
    return render_rows("Latest succeeded results", rows, "Latest succeeded result rows.")


@app.route("/ip_errors_since_success")
def ip_errors_since_success():
    rows = get_ip_errors_since_success(ROW_LIMIT)
    chart_rows = build_ip_error_distribution(rows)
    return render_template(
        "ip_errors_since_success.html",
        title="Most errors per IP since last success",
        description="Failure counts after each IP address last succeeded.",
        rows=rows,
        row_limit=ROW_LIMIT,
        chart_rows=chart_rows,
        chart_y_ticks=build_ip_error_y_ticks(chart_rows),
    )


@app.route("/site_summary")
def site_summary():
    rows = query_all(
        LATEST_RUN_CTE
        + """
        SELECT
            Site_Name,
            COUNT(*) AS Total,
            COALESCE(SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END), 0) AS Success,
            COALESCE(SUM(CASE WHEN Is_Success = 0 THEN 1 ELSE 0 END), 0) AS Failed,
            COALESCE(ROUND(
                100.0 * SUM(CASE WHEN Is_Success = 1 THEN 1 ELSE 0 END) /
                NULLIF(COUNT(*), 0),
                1
            ), 0) AS Success_Rate
        FROM latest_run
        GROUP BY Site_Name
        ORDER BY Failed DESC, Total DESC
        LIMIT ?;
        """,
        (ROW_LIMIT,),
    )
    return render_rows("Site coverage mapping", rows, "Grouped by site name from the latest run.")


@app.route("/search_ip")
def search_ip():
    search_text = (request.args.get("q") or request.args.get("ip") or "").strip()

    if not search_text:
        return render_rows(
            "Search results",
            [],
            "Type an IP address, site name, or detail text in the search box.",
        )

    like_text = f"%{search_text}%"
    rows = query_all(
        """
        SELECT ID, IP_Address, Port, Site_Name, Is_Success, Details, "Date"
        FROM tbl_Results
        WHERE
            IP_Address LIKE ?
            OR Site_Name LIKE ?
            OR Details LIKE ?
        ORDER BY "Date" DESC
        LIMIT ?;
        """,
        (like_text, like_text, like_text, ROW_LIMIT),
    )
    return render_rows(f"Search: {search_text}", rows)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
